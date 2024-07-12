import io
import datetime
import logging
from typing import List, Union, Tuple, Dict
from openpyxl import load_workbook

"""
Skills Data Operations Class

Pre-reqs:
    1. Install the following environment variables prior to execution:
        a. azure_blob_connection_string
        b. azure_blob_container_name
        
        Tip: Include these environment variables in your environment setup to automate them upon execution.
"""

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class SkillsDataOperations:
    '''
    Skills Matrix Data Operations

    Methods:
    __init__(self) -> None: Initializes the SkillsDataOperations class.
    get_column_headers(self, worksheet) -> List[str]: Processes column headers.
    get_row_values(self, worksheet) -> List[List[Union[str, datetime.datetime]]]: Processes row values.
    survey_data_preparation(self, blob_obj: bytes) -> Tuple[List[str], List[List[Union[str, datetime.datetime]]]]: Prepares data from an Excel file.
    get_language_value(self, header_value: str, value: str, zero_list: List[str]) -> str: Converts a language value to a customized string.
    stage_row_values(self, header_vals: List[str], row_vals: List[List[Union[str, datetime.datetime]]], zero_list: List[str]) -> List[Dict[str, List[str]]]: Stages data for insertion into a Tableau data source.
    nested_lists_to_csv_strings(self, data: List[Dict[str, List[str]]]) -> List[Dict[str, str]]: Converts nested lists into comma-separated strings.
    map_work_experience_values(self, data: List[Dict[str, Union[List[str], str]]], paired_values: Dict[str, int]) -> List[Dict[str, str]]: Maps work experience descriptors to integer values.
    '''

    def __init__(self) -> None:
        logging.info("Initializing SkillsDataOperations class.")
        pass

    def get_column_headers(self, worksheet) -> List[str]:
        """
        Processes column headers.

        Args:
            worksheet: The Excel worksheet object.

        Returns:
            List[str]: A list of formatted column headers.
        """
        logging.info("Processing column headers.")
        try:
            headers = [cell.value.replace('\u00a0', ' ') for cell in worksheet[1]]
            logging.info(f"Successfully processed {len(headers)} column headers.")
            return headers
        except Exception as e:
            logging.error(f"Error processing column headers: {e}")
            raise

    def get_row_values(self, worksheet) -> List[List[Union[str, datetime.datetime]]]:
        """
        Processes row values.

        Args:
            worksheet: The Excel worksheet object.

        Returns:
            List[List[Union[str, datetime.datetime]]]: A list of formatted row values.
        """
        logging.info("Processing row values.")
        try:
            row_vals = [
                [
                    str(cell.value).strip() if not isinstance(cell.value, datetime.datetime) else cell.value.strftime('%Y-%m-%d')
                    for cell in row
                ]
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column)
            ]
            logging.info(f"Successfully processed {len(row_vals)} rows of data.")
            return row_vals
        except Exception as e:
            logging.error(f"Error processing row values: {e}")
            raise

    def survey_data_preparation(self, blob_obj: bytes) -> Tuple[List[str], List[List[Union[str, datetime.datetime]]]]:
        """
        Prepares data from an Excel file.

        Args:
            blob_obj (bytes): A bytes object containing the Excel file.

        Returns:
            Tuple[List[str], List[List[Union[str, datetime.datetime]]]]: Column headers and row values.

        Raises:
            ValueError: An error occurred when reading the Excel file.
        """
        logging.info("Preparing survey data from blob object.")
        try:
            workbook = load_workbook(io.BytesIO(blob_obj), read_only=True)
            worksheet = workbook.active
        except Exception as e:
            logging.error(f"Error when reading Excel: {e}")
            raise ValueError(f'Error when reading Excel: {e}')
        
        col_headers = self.get_column_headers(worksheet)
        row_vals = self.get_row_values(worksheet)
        logging.info("Survey data preparation completed.")
        return col_headers, row_vals

    def get_language_value(self, header_value: str, value: str, zero_list: List[str]) -> str:
        """
        Converts a language value to a customized string.

        Args:
            header_value (str): Title of sufficient value.
            value (str): The value to convert to binary.
            zero_list (List[str]): A list of values considered as zero.

        Returns:
            str: A customized string value.
        """
        try:
            if value.strip() in {'2', '3', '4', '5'}:
                return f'{header_value} (Level {value})'
            elif value not in zero_list:
                return value
            return ""
        except Exception as e:
            logging.error(f"Error converting language value: {e}")
            raise

    def stage_row_values(self, header_vals: List[str], row_vals: List[List[Union[str, datetime.datetime]]], 
                         zero_list: List[str] = None) -> List[Dict[str, List[str]]]:
        """
        Stages data for insertion into a Tableau data source.

        Args:
            header_vals (List[str]): Column headers.
            row_vals (List[List[Union[str, datetime.datetime]]]): Row values.
            zero_list (List[str]): Values to be ignored.

        Returns:
            List[Dict[str, List[str]]]: Staged data.
        """
        logging.info("Staging row values for Tableau data source.")
        if zero_list is None:
            zero_list = ['N', 'No', 'NO', 'Np', 'no', 'n', 'noo', 'nm', 'none', 'None', 'NOne', 'nOne', ' ', '', 'null', None]
        
        try:
            special_col_indexes = {
                'work_experience': header_vals.index('Work Experience'),
                'space': header_vals.index('Space'),
                'aircraft_power': header_vals.index('Aircraft Power Generation'),
                'other_standards': header_vals.index('Other Standards'),
                'devsecops': header_vals.index('DevSecOps'),
                'other_languages': header_vals.index('Other Languages'),
                'other_tools': header_vals.index('Other Tools')
            }

            staged_row_values = []
            for row in row_vals:
                new_row = {
                    'Basic Information': [],
                    'Industry Experience': [],
                    'Domains': [],
                    'Standards': [],
                    'Skills': [],
                    'Languages': [],
                    'Tools': []
                }

                for i, value in enumerate(row):
                    if i <= special_col_indexes['work_experience']:
                        new_row['Basic Information'].append(value)
                    elif i <= special_col_indexes['space'] and value.strip().lower() == 'yes':
                        new_row['Industry Experience'].append(header_vals[i])
                    elif i <= special_col_indexes['aircraft_power'] and value.strip().lower() == 'yes':
                        new_row['Domains'].append(header_vals[i])
                    elif i < special_col_indexes['other_standards'] and value.strip().lower() == 'yes':
                        new_row['Standards'].append(header_vals[i])
                    elif i == special_col_indexes['other_standards'] and value not in zero_list:
                        new_row['Standards'].append(value)
                    elif i <= special_col_indexes['devsecops'] and value.strip().lower() == 'yes':
                        new_row['Skills'].append(header_vals[i])
                    elif i < special_col_indexes['other_languages'] and value in {'2', '3', '4', '5'}:
                        language_value = self.get_language_value(header_vals[i], value, zero_list)
                        if language_value:
                            new_row['Languages'].append(language_value)
                    elif i == special_col_indexes['other_languages'] and value not in zero_list:
                        new_row['Languages'].append(value)
                    elif i < special_col_indexes['other_tools'] and value.strip().lower() == 'yes':
                        new_row['Tools'].append(header_vals[i])
                    elif i == special_col_indexes['other_tools'] and value not in zero_list:
                        new_row['Tools'].append(value)

                staged_row_values.append(new_row)
            logging.info("Staging of row values completed.")
            return staged_row_values
        except Exception as e:
            logging.error(f"Error staging row values: {e}")
            raise

    def nested_lists_to_csv_strings(self, data: List[Dict[str, List[str]]]) -> List[Dict[str, str]]:
        """
        Converts nested lists into comma-separated strings.

        Args:
            data (List[Dict[str, List[str]]]): Skills data.

        Returns:
            List[Dict[str, str]]: Skills data with CSV strings.
        """
        logging.info("Converting nested lists to CSV strings.")
        try:
            for entry in data:
                for key in entry:
                    if key != 'Basic Information':
                        entry[key] = ', '.join(entry[key])
            logging.info("Conversion to CSV strings completed.")
            return data
        except Exception as e:
            logging.error(f"Error converting nested lists to CSV strings: {e}")
            raise

    def map_work_experience_values(self, data: List[Dict[str, Union[List[str], str]]], 
                                   paired_values: Dict[str, int] = None) -> List[Dict[str, str]]:
        """
        Maps work experience descriptors to integer values.

        Args:
            data (List[Dict[str, Union[List[str], str]]]): Skills data.
            paired_values (dict): Mapping from string descriptors to integer values.

        Returns:
            List[Dict[str, str]]: Updated skills data with mapped work experience values.
        """
        logging.info("Mapping work experience values.")
        if paired_values is None:
            paired_values = {
                "0 to 4": 1, "5 to 9": 2, "10 to 14": 3, 
                "15 to 19": 4, "20 to 24": 5, "25 to 29": 6, 
                "30 or more": 7
            }
        
        try:
            for candidate in data:
                experience_value = candidate['Basic Information'][8]
                candidate['Basic Information'][8] = paired_values.get(experience_value, experience_value)
            logging.info("Mapping of work experience values completed.")
            return data
        except Exception as e:
            logging.error(f"Error mapping work experience values: {e}")
            raise